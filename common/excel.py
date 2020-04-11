import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.styles.colors import BLACK, WHITE

from common.attrs import get_attr_value_name
from common.utils import trans_to_chinese_code


# 导出EXCEL
class ExportExcelMixin(object):

    # 导出Excel
    def export_as_excel(self, request, queryset):
        # 文件名
        filename = trans_to_chinese_code(self.model._meta.verbose_name)
        response = HttpResponse(content_type='application/msexcel')
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        wb = Workbook()
        ws = wb.active
        # 设置Excel列名
        ws.append(self.get_fields_chinese_name())
        # 设置表头样式
        self.set_header_style(ws)
        # 添加行
        index = 1
        for obj in queryset:
            field_names = self.get_fields_name()
            for field_name in field_names:
                data = [f'{self.get_field_value(obj, field_name)}' for field_name in field_names]
                if data is None:
                    data = ''
            row = ws.append(data)
            index = index + 1
            self.set_row_style(ws, index)  # 设置每一行的格式

        wb.save(response)
        return response

    export_as_excel.short_description = '导出Excel'

    # 获取展示列
    def get_fields_name(self):
        return list(self.list_display)

    # 获取展示列的中文名
    def get_fields_chinese_name(self):
        field_names = []

        for field_name in self.get_fields_name():
            field = getattr(self.model, field_name)
            if not field:
                raise Exception('展示列表list_dispay中%s字段识别不到!' % field_name)
            verbose_name = ''
            if field.field and field.field.verbose_name:
                verbose_name = field.field.verbose_name
            field_names.append(verbose_name)
        return field_names

    # 获取字段值
    def get_field_value(self, obj, field_name):
        value = getattr(obj, field_name)
        if value is None:
            value = ''
        else:
            # 查找是否有主数据定义
            chinese = get_attr_value_name(self.model._meta.object_name, field_name, value)
            if chinese:
                value = chinese
        return value

    # 设置Excel列头样式
    def set_header_style(self, ws, width=None):
        font = Font(u'宋体', size=12, bold=True, italic=False, strike=False, color=WHITE)
        row = ws[1]
        fill = PatternFill("solid", fgColor="409eff")  # 背景色
        for cell in row:
            cell.font = font  # 设置字体样式
            self.set_border(cell)  # 设置边框
            self.set_align(cell)  # 设置对齐方式
            cell.fill = fill  # 设置背景色
            if width is None:
                width = 30  # 列宽默认为30
            ws.column_dimensions[cell.column_letter].width = 30  # 设置列的宽度

    # 设置Excel每一行的样式
    def set_row_style(self, ws, index, font=None):
        row = ws[index]
        if font is None:
            font = Font(u'宋体', size=10, bold=False, italic=False, strike=False, color=BLACK)
        for cell in row:
            cell.font = font  # 设置字体样式
            self.set_border(cell)  # 设置边框
            self.set_align(cell)  # 设置对齐方式

    # 设置Excel每一列的样式
    # def set_col_style(self, ws, col_name, font=None):
    #     col = ws.column_dimensions[col_name]
    #     if font is None:
    #         font = Font(u'宋体', size=11, bold=False, italic=False, strike=False, color='000000')
    #     col.font = font
    #
    # # 设置Excel每一单元格的样式
    # def set_cell_style(self, ws, col_index, row_index, font=None):
    #     if font is None:
    #         font = Font(u'宋体', size=11, bold=False, italic=False, strike=False, color='000000')
    #     ws.cell(row=row_index, column=col_index).font = font

    # 边框设置
    def set_border(self, obj, side=None):
        if side is None:
            side = Side(border_style='thin',
                        color='FF000000')
        border = Border(left=side,
                        right=side,
                        top=side,
                        bottom=side,
                        diagonal=side,
                        diagonal_direction=0,
                        outline=side,
                        vertical=side,
                        horizontal=side
                        )
        obj.border = border

    # 对齐
    def set_align(self, obj, align=None):
        if align is None:
            align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        obj.alignment = align


# 导入EXCEL
class ImportExcelMixin(object):
    def import_excel(self, filename, sheetname):
        workbook = openpyxl.load_workbook(filename)
        # 通过文件名得到文件对象
        sheet_name = workbook.get_sheet_by_name(sheetname)
        # 通过名称得到工作簿对象
        # rows_sheet = sheet_name.rows
        # 按行生成工作表中所有单元格对象，生成器类型
        rows = [item.value for item in list(sheet_name.rows)[1]]
        print(rows)
        # 第二行的内容
        cols = [item.value for item in list(sheet_name.columns)[1]]
        print(cols)
        # 第二列的内容
        rows_sheet = sheet_name.iter_rows()
        for item in rows_sheet:
            for call in item:
                print(call.coordinate, call.value)
        # 遍历所有内容
        cell_1_2 = sheet_name.cell(row=1, column=2).value
        print(cell_1_2)
        # 查看第一行第二列的单元格内容
        print(sheet_name.max_row, sheet_name.max_column)
